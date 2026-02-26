import logging
import os
import pickle
import traceback
import numpy as np
import pandas as pd
from pathlib import Path
from Custom_Exception import CustomException
from threading import Thread
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation
from tkinter import messagebox
import gc

# log_file_path = "C:/Ericsson_Application_Logs/CLI_Automation_Logs/"
# Path(log_file_path).mkdir(parents=True,exist_ok=True)
# log_file = os.path.join(log_file_path,"Sheet_Creation_Task.log")
# logging.basicConfig(filename=log_file,
#                              filemode="a",
#                              format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: ({'%(module)s'}): {'%(message)s'}",
#                              datefmt='%d-%b-%Y %I:%M:%S %p',
#                              encoding= "UTF-8",
#                              level=logging.DEBUG)
# logging.captureWarnings(capture=True)

flag = ''
file_reader = ''


def pickle_checker(host_details_df: pd.DataFrame) -> list:
    """
        Checks the uploaded workbook with the pre-defined pickle in User/{signum}/AppData/Local/CLI_Automation/Host_details_Pickle_file/Host_detail_pickle.pkl
        
        Arguments : (host_details_df)
            host_details_df : pandas.DataFrame
                description =====> dataframe containing details from the uploaded Host Details workbook
            
        return result             : list
                description =====> list containing the status and message to show as exception in the main_func where 
                                    result[0] ===> status
                                    result[1] ===> message:str
                                                    or
                                                   None
    """

    logging.basicConfig(filename=log_file,
                        filemode="a",
                        format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: (Main-application/{'%(module)s'}/pickle_checker): {'%(message)s'}",
                        datefmt='%d-%b-%Y %I:%M:%S %p',
                        encoding="UTF-8",
                        level=logging.DEBUG)
    logging.captureWarnings(capture=True)

    username = os.popen(r'cmd.exe /C "echo %username%"').read()
    logging.info(r"Got the username from the command \'cmd.exe /C \"echo %username%\"\'")
    path_for_the_host_details_pickle = rf"C:\Users\{username.strip()}\AppData\Local\CLI_Automation\Host_details_Pickle_file\Host_details.pkl"
    Path(os.path.dirname(path_for_the_host_details_pickle)).mkdir(parents=True, exist_ok=True)
    result = []

    logging.info(f"Checking if the there is any pickle file existing at \'{path_for_the_host_details_pickle}\'")
    try:
        if not Path(path_for_the_host_details_pickle).exists():
            logging.info(
                "Didn't find the pickle file for Host Details so creating the pickle file for the Host Details\n")

            host_details_df.to_pickle(path=path_for_the_host_details_pickle,
                                      compression=None,
                                      protocol=pickle.HIGHEST_PROTOCOL)

            result = ['Successful', None]

        else:
            logging.info("Found the pickle file for the Host Details")
            creation_time_of_host_details_pickle_file = datetime.fromtimestamp(
                os.path.getmtime(path_for_the_host_details_pickle))

            today = datetime.now()
            timedelta_var = today - creation_time_of_host_details_pickle_file
            logging.debug("Finding the modified time in seconds")
            timedelta_var_hour = (timedelta_var.days * 24 + timedelta_var.seconds)//3600

            if (timedelta_var.days >= 1) or (timedelta_var_hour >= 20):
                logging.debug(f"{timedelta_var.days= } days")
                logging.debug(f"{timedelta_var_hour= } hours")
                logging.info(
                    "pickle file for Host Details found is older so creating the pickle file for the Host Details\n")
                os.remove(path_for_the_host_details_pickle)

                host_details_df.to_pickle(path=path_for_the_host_details_pickle,
                                          compression=None,
                                          protocol=pickle.HIGHEST_PROTOCOL)

                result = ['Successful', None]

            else:
                logging.info("Loading the pickle and checking the file with the pickle\n")

                pickle_to_df_of_previous_host_details = pd.read_pickle(
                    filepath_or_buffer=path_for_the_host_details_pickle,
                    compression=None)
                logging.info(f"Read the df from pickle ==> \n{pickle_to_df_of_previous_host_details.to_markdown()}\n")
                logging.debug(f"{timedelta_var.days= } days")
                logging.debug(f"{timedelta_var_hour= } hours")

                if len(pickle_to_df_of_previous_host_details['Host_IP']) == len(host_details_df['Host_IP']):
                    temp_result_df = pickle_to_df_of_previous_host_details['Host_IP'].compare(
                        host_details_df['Host_IP'])

                    logging.debug(
                        f"comparison between the uploaded host details and pickled host_details ==>\n{temp_result_df.to_markdown()}\n")

                    if len(temp_result_df) > 0:
                        messagebox.showinfo("Updated Host Details Information",
                                              f"New Host IP/s entry found in latest uploaded \'Host Details\', Please Cross-Check before proceeding further!\nNew Host IPs :- {', '.join(str(element) for element in temp_result_df['other'])}")
                        host_details_df.to_pickle(path=path_for_the_host_details_pickle,
                                                  compression=None,
                                                  protocol=pickle.HIGHEST_PROTOCOL)

                    else:
                        messagebox.showinfo("Updated Host Details Information",
                                              "You have uploaded same \'Host Details\'!")

                    result = ['Successful', None]

                else:
                    previous_host_details_pickle_to_df_host_ip = pickle_to_df_of_previous_host_details[
                        'Host_IP'].to_numpy()
                    uploaded_host_details_df_host_ip = host_details_df['Host_IP'].to_numpy()

                    if previous_host_details_pickle_to_df_host_ip.size < uploaded_host_details_df_host_ip.size:
                        temp_delta = np.setdiff1d(ar1=uploaded_host_details_df_host_ip,
                                                  ar2=previous_host_details_pickle_to_df_host_ip)

                        messagebox.showinfo("Updated Host Details Information",
                                              f"New Host IP/s entry found in latest uploaded \'Host Details\', Please Cross-Check before proceeding further!\nNew Host IPs :- {', '.join(temp_delta.astype(dtype=str))}")
                        host_details_df.to_pickle(path=path_for_the_host_details_pickle,
                                                  compression=None,
                                                  protocol=pickle.HIGHEST_PROTOCOL)

                    else:
                        temp_delta = np.setdiff1d(ar1=previous_host_details_pickle_to_df_host_ip,
                                                  ar2=uploaded_host_details_df_host_ip)

                        messagebox.showinfo("Updated Host Details Information",
                                              f"Few Host IPs entry have been removed from latest uploaded \'Host Details\', Please Cross-Check before proceeding further!\nRemoved Host IPs :- {', '.join(temp_delta.astype(dtype=str))}")
                        host_details_df.to_pickle(path=path_for_the_host_details_pickle,
                                                  compression=None,
                                                  protocol=pickle.HIGHEST_PROTOCOL)

                    result = ['Successful', None]

    except Exception as e:
        logging.error(f"Exception Occurred ======>\n{traceback.format_exc()}\n\n{e}\n")
        result = ['Unsuccessful', e.args[1]]

    finally:
        logging.info(f"Returning the value {result}")
        return result


def mpbn_node_login_file_creater(**kwargs) -> None:
    """
        Creates MPBN_Node_Login workbook for entering host details
        
        Arguments : (**kwargs)
            **kwargs ===>   parent_dir : str
                                description =====> contains the path for the parent directory for the selected Host Details workbook
                            
                            host_details_file_df : pandas.DataFrame
                                description =====> contains the dataframe object from reading the 'Host Details' file uploaded by user

        returns None
    """

    logging.basicConfig(filename=log_file,
                        filemode="a",
                        format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: (Main-application/{'%(module)s'}/mpbn_node_login_file_creater): {'%(message)s'}",
                        datefmt='%d-%b-%Y %I:%M:%S %p',
                        encoding="UTF-8",
                        level=logging.DEBUG)
    logging.captureWarnings(capture=True)

    host_details_file_df = kwargs['host_details_file_df']
    parent_dir = kwargs['parent_dir']
    path_for_user_mpbn_activity_nodes = os.path.join(parent_dir, "MPBN_Node_Login.xlsx")
    logging.debug(
        f"Created the path for creation of \'MPBN_Node_Login.xlsx\' ==> {path_for_user_mpbn_activity_nodes}\n")

    if not Path(path_for_user_mpbn_activity_nodes).exists():
        logging.debug(f"\'MPBN_Node_Login.xlsx\'File not found, so creating the file")
        wkbk = Workbook()
        wkbk.save(path_for_user_mpbn_activity_nodes)
        wkbk.close()
        del wkbk

    columns_for_mpbn_node_login_sheet = ["S.N.",
                                         "Host Name",
                                         "Host IP",
                                         "Vendor Details",
                                         "Host User Name",
                                         "Host Password",
                                         "OSS IP",
                                         "OSS User Name",
                                         "OSS Password",
                                         "CR ID",
                                         "Login Port For Node",
                                         "Secondary Path IP",
                                         "Secondary Path User Name",
                                         "Secondary Path Password",
                                         "CLI Execution File Required"]

    dataframe_for_mpbn_node_login = pd.DataFrame(columns=columns_for_mpbn_node_login_sheet)

    logging.debug("Entering details of the host details in \'MPBN_Node_Login.xlsx\'")

    dataframe_for_mpbn_node_login['Host Name'] = host_details_file_df['Host_Name']
    dataframe_for_mpbn_node_login['Host IP'] = host_details_file_df['Host_IP']
    dataframe_for_mpbn_node_login['CR ID'] = [f'{row['Vendor']}_{row['CR_ID']}' for _, row in
                                              host_details_file_df.iterrows()]
    dataframe_for_mpbn_node_login['Vendor Details'] = [element.strip() if (element.strip() != 'Cisco') else '' for
                                                       element in host_details_file_df['Vendor']]
    dataframe_for_mpbn_node_login['S.N.'] = dataframe_for_mpbn_node_login.index + 1
    dataframe_for_mpbn_node_login['OSS IP'] = '150.236.9.9'
    dataframe_for_mpbn_node_login['Login Port For Node'] = 22

    writer = pd.ExcelWriter(path_for_user_mpbn_activity_nodes, mode='w', engine='openpyxl')
    dataframe_for_mpbn_node_login.to_excel(writer, sheet_name='Node_Login', index=False)

    logging.debug("Created the file \'MPBN_Node_Login.xlsx\' with sheetname \'Node_Login\'\n")
    writer.close()
    del writer

    # ensuring no other sheet exists in mpbn node login, than Node_Login
    wkbk = load_workbook(path_for_user_mpbn_activity_nodes)
    sheetnames = wkbk.sheetnames

    if (len(sheetnames) > 1):
        for sheetname in sheetnames:
            if (sheetname != 'Node_Login'):
                del wkbk[sheetname]

    ws = wkbk['Node_Login']

    col_widths = []
    for row_values in ws.iter_rows(values_only=True):
        for j, value in enumerate(row_values):
            if len(col_widths) > j:
                if col_widths[j] < len(str(value)):
                    col_widths[j] = len(str(value))
            else:
                col_widths.insert(j, len(str(value)))
    logging.debug(f"Got the column widths for the headers =>\n{col_widths}")

    # Standardising the length of each column in the sheet.

    i = 1
    while (i <= len(col_widths)):
        column_width = col_widths[i - 1]
        if column_width <= 47:
            ws.column_dimensions[get_column_letter(i)].width = column_width + 3
        else:
            ws.column_dimensions[get_column_letter(i)].width = 50

        i += 1

    logging.debug("Standardised the length of each column in the sheet")

    # for column in range(1,ws.max_column+1):   # ws.max_column returns the total number of columns present
    font_style = Font(bold=True, color=Color(rgb='00000000'))
    i = 1
    while (i <= ws.max_column):
        column = i
        col = get_column_letter(column)
        color_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[col + '1'].font = font_style
        ws[col + '1'].fill = color_fill
        ws[col + '1'].alignment = Alignment(horizontal='center', vertical='center')

        i += 1

    for row in ws:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            cell.border = Border(top=Side(border_style='medium', color='000000'),
                                 bottom=Side(border_style='medium', color='000000'),
                                 left=Side(border_style='medium', color='000000'),
                                 right=Side(border_style='medium', color='000000'))
    maxrows = ws.max_row

    # Creating a list for the selection.
    vendor_list = ["Cisco_ASA_Firewall",
                   "Cisco_IOS_Router",
                   "Cisco_IOS_Switch",
                   "Cisco_Nexus",
                   "Cisco_XR",
                   "Ericsson",
                   "Extreme",
                   "Huawei",
                   "Huawei_Firewall",
                   "Juniper",
                   "Nokia",
                   "NWI"]

    vendor_list = f"{','.join(vendor_list)}"
    vendor_list = f'"{vendor_list}"'

    Yes_no_list = ["Yes", "No"]
    Yes_no_list = f"{','.join(Yes_no_list)}"
    Yes_no_list = f'"{Yes_no_list}"'

    # Defining the rule
    rule1 = DataValidation(type="list", formula1=vendor_list, allow_blank=False)
    rule1.error = "Your Entry is invalid"
    rule1.errorTitle = "Invalid Entry"

    rule1.prompt = "Please Select Vendor from the list"
    rule1.promptTitle = "List Selection"

    # Adding the rule to the worksheet
    ws.add_data_validation(rule1)

    # Defining the range for the rule to work in the worksheet.
    range_setter = f"D2:D{maxrows}"

    # Adding the range to the rule
    rule1.add(range_setter)

    rule2 = DataValidation(type="list", formula1=Yes_no_list, allow_blank=False)
    rule2.error = "Your Entry is invalid"
    rule2.errorTitle = "Invalid Entry"

    rule2.prompt = "Please Select Yes/No from the list"
    rule2.promptTitle = "List Selection"

    # Adding the rule to the worksheet
    ws.add_data_validation(rule2)

    # Defining the range for the rule to work in the worksheet.
    neo_range_setter = f"O2:O{maxrows}"

    # Adding the range to the rule
    rule2.add(neo_range_setter)

    wkbk.save(path_for_user_mpbn_activity_nodes)
    wkbk.close()
    del wkbk


def file_path_saver(*args) -> None:
    """Saves the file path of host details
    """
    username = (os.popen(r'cmd.exe /C "echo %username%"').read()).strip()

    path_to_save_file_name_path = f"C:\\Users\\{username}\\AppData\\Local\\CLI_Automation\\host_details_file_path.txt"
    Path(os.path.dirname(path_to_save_file_name_path)).mkdir(parents=True, exist_ok=True)

    file_name = args[0]
    with open(path_to_save_file_name_path, 'w') as _f:
        _f.write(file_name)
        _f.close()

    del _f


def sheet_creater(**kwargs) -> str:
    # logging.basicConfig(filename=log_file,
    #                     filemode="a",
    #                     format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: (Main-application/{'%(module)s'}/sheet_creater): {'%(message)s'}",
    #                     datefmt='%d-%b-%Y %I:%M:%S %p',
    #                     encoding="UTF-8",
    #                     level=logging.DEBUG)
    # logging.captureWarnings(capture=True)

    try:
        logging.info(f"Starting the Sheet Creater for {os.path.basename(kwargs['file'])}")
        host_ips_sheets_required = kwargs['host_ips_sheets_required']
        logging.info(f"{host_ips_sheets_required =}")

        file = kwargs['file']
        logging.info(f"{file =}")

        standard_input = kwargs['standard_design_template_path']
        logging.info(f"{standard_input = }")

        file_name = os.path.basename(file)
        standard_input_file_name = os.path.basename(standard_input)

        logging.debug(
            f"Reading the '{standard_input_file_name}' for the 'Standard Template' worksheet using 'openpyxl'")
        standard_input_workbook = load_workbook(standard_input)
        standard_input_template_worksheet = standard_input_workbook['Standard Template']
        max_rows = standard_input_template_worksheet.max_row
        max_columns = standard_input_template_worksheet.max_column

        logging.debug(f"Reading the {file_name} using openpyxl")
        host_ips_sheets_required_workbook = load_workbook(file)
        # print(type(host_ips_sheets_required_workbook))

        host_ips_sheetnames_present_in_the_workbook = host_ips_sheets_required_workbook.sheetnames
        f = pd.ExcelFile(file)
        # print("sheetnames for file",file,"===>",f.sheet_names)
        f.close()
        del f

        # print(file,"===>",host_ips_sheetnames_present_in_the_workbook)
        logging.debug(f"Sheetnames present in {file_name} ====> {host_ips_sheetnames_present_in_the_workbook}")

        host_ips_sheetnames_present_in_the_workbook = np.array(host_ips_sheetnames_present_in_the_workbook)
        logging.debug(
            f"Host IPs mentioned in uploaded 'Host Details' workbook for filename {file_name}=> {host_ips_sheets_required}")

        setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array = np.setdiff1d(
            ar1=host_ips_sheetnames_present_in_the_workbook,
            ar2=host_ips_sheets_required)

        # setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array = setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array[np.where(setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array != 'Sheet')]
        setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array = \
            setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array[
                ~np.char.startswith(setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array,
                                    'Sheet')]
        logging.debug(
            f"Got the setdiff_between_workbook_and_host_ips_of_uploaded_host_details_array for '{file_name}'==>\n{setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array}")

        if setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array.size > 0:
            logging.debug(
                f"Raising CustomException as setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array.size > 0\nfor\n{setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array}")
            raise CustomException("Host IPs Mismatch",
                                  f"Host IP mismatch found in uploaded \'Host Details\' and existing {file_name}\nMismatch Host IPs: {', '.join(setdiff_between_input_design_workbook_and_host_ips_of_uploaded_host_details_array.astype(str))}")

        logging.debug("Getting the length for widths for the columns of the worksheet")
        width_list = []
        i = 1
        while i <= max_rows:
            j = 1
            while j <= max_columns:
                cell_selected = f'{get_column_letter(j)}{i}'
                if len(width_list) < j:
                    width_list.insert(j - 1, len(str(standard_input_template_worksheet[cell_selected].value)) + 3)

                else:
                    if (len(str(standard_input_template_worksheet[cell_selected].value)) + 3) > width_list[j - 1]:
                        width_list[j - 1] = len(str(standard_input_template_worksheet[cell_selected].value)) + 3
                j += 1
            i += 1

        # print(host_ips_sheetnames_present_in_the_workbook)

        i = 0
        while i < host_ips_sheets_required.size:
            logging.debug(f"Checking for presence of {host_ips_sheets_required[i]} in {file_name} \n")
            logging.debug(
                f"Checking the condition for {host_ips_sheets_required[i] not in host_ips_sheetnames_present_in_the_workbook}")
            if host_ips_sheets_required[i] not in host_ips_sheetnames_present_in_the_workbook:

                logging.debug(f"Creating {host_ips_sheets_required[i]} sheet in {file_name}")
                worksheet = host_ips_sheets_required_workbook.create_sheet(title=host_ips_sheets_required[i])

                j = 1
                while j <= max_rows:
                    k = 1
                    while k <= max_columns:
                        cell_selected = f'{get_column_letter(k)}{j}'
                        # color = standard_input_template_worksheet[cell_selected].fill.start_color
                        side = Side(color=Color(rgb='00000000'), style='thin')
                        # logging.debug(f"Writing {cell_selected} for {host_ips_sheets_required[i]} in {file_name}")

                        worksheet[cell_selected].value = standard_input_template_worksheet[cell_selected].value
                        # with open(r"C:\Users\emaienj\Downloads\VPLS_CLI_Design_Documents\VPLS_CLI_Design_Documents\demo.txt","a+") as f:
                        #     f.write(f"color of {get_column_letter(k)}{j} =====> {color}\n\n")
                        # if(not ( (color.rgb == None) or (color == Color(rgb = '00000000')) ) ):
                        #     # font.bold = True
                        #     worksheet[cell_selected].fill = PatternFill(start_color = color,
                        #                                                 end_color = color,
                        #                                                 bgColor = color,
                        #                                                 fill_type = 'solid')

                        # worksheet[cell_selected].font = 
                        # worksheet[cell_selected].alignment = Alignment(horizontal='center',
                        #                                                vertical='center')

                        worksheet[cell_selected].font = standard_input_template_worksheet[cell_selected].font.copy()
                        worksheet[cell_selected].alignment = standard_input_template_worksheet[
                            cell_selected].alignment.copy()
                        worksheet[cell_selected].fill = standard_input_template_worksheet[cell_selected].fill.copy()
                        worksheet[cell_selected].border = Border(left=side,
                                                                 right=side,
                                                                 top=side,
                                                                 bottom=side)

                        k += 1
                    j += 1

                j = 1
                while (j <= max_columns):
                    worksheet.column_dimensions[get_column_letter(j)].width = width_list[j - 1]
                    j += 1
            i += 1

        logging.info("Removing the extra sheet that are not required in the workbook")
        for sheet in host_ips_sheetnames_present_in_the_workbook:
            if sheet.startswith('Sheet'):
                del host_ips_sheets_required_workbook[sheet]

        logging.info(f"Saving the file {file_name}")
        host_ips_sheets_required_workbook.save(file)

        host_ips_sheets_required_workbook.close()
        del host_ips_sheets_required_workbook

        logging.debug(f"Closing the {standard_input_file_name}")
        standard_input_workbook.close()
        del standard_input_workbook

    except CustomException as e:
        global flag
        flag = 'Unsuccessful'
        logging.error(
            f"{traceback.format_exc()}\n\nraised CustomException==>\ntitle = {e.title}\nmessage = {e.message}")

    except Exception as e:
        flag = 'Unsuccessful'
        logging.error(f"{traceback.format_exc()}\n\nException:==>{e}")
        messagebox.showerror("Exception Occurred!", str(e))


def file_creater(**kwargs) -> None:
    logging.basicConfig(filename=log_file,
                        filemode="a",
                        format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: (Main-application/{'%(module)s'}/file_creater): {'%(message)s'}",
                        datefmt='%d-%b-%Y %I:%M:%S %p',
                        encoding="UTF-8",
                        level=logging.DEBUG)
    logging.captureWarnings(capture=True)
    # print("running file creater")
    file = kwargs['file']
    wb = Workbook()
    wb.save(file)
    wb.close()
    del wb


# kwargs.keys() ==> (file_name)
def main_func(**kwargs) -> str:
    """
        Main Function for the sheet creater, to create Sheets from Vendor Specific Design Input Templates
        
        Arguments : (**kwargs)
            **kwargs ===> file_name : str
                            description =====> contains the path of the uploaded host details file.
        
        return flag
            flag : str
                description =====> contains 'Unsuccessful' or 'Successful' string corresponding the status of execution completion
    """
    log_file_path = "C:/Ericsson_Application_Logs/CLI_Automation_Logs/"
    Path(log_file_path).mkdir(parents=True, exist_ok=True)
    global log_file, file_reader
    log_file = os.path.join(log_file_path, "Sheet_Creation_Task.log")

    today = datetime.now()
    today = today.replace(hour=0, minute=0, second=0)

    if os.path.exists(log_file):
        # getting the creation time of the log file
        log_file_create_time = datetime.fromtimestamp(os.path.getctime(log_file))

        if log_file_create_time < today:
            os.remove(log_file)

    logging.basicConfig(filename=log_file,
                        filemode="a",
                        format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: (Main-application/{'%(module)s'}/main_func): {'%(message)s'}",
                        datefmt='%d-%b-%Y %I:%M:%S %p',
                        encoding="UTF-8",
                        level=logging.DEBUG)
    logging.captureWarnings(capture=True)

    global flag
    flag = ''

    logging.info(
        "#############################################<<Starting the Sheet Creation Task>>#################################################")

    host_details_file_name = kwargs['file_name']
    file_reader = pd.ExcelFile(host_details_file_name, engine='openpyxl')

    try:

        host_details_file_sheetnames = file_reader.sheet_names

        # Checking for the sheet 'Host Details' in the uploaded workbook.
        logging.debug("Checking for the 'Host Details' worksheet in the uploaded workbook")
        if "Host Details" not in host_details_file_sheetnames:
            raise CustomException("Sheet Not Found!",
                                  "'Host Details' workbook not found in the uploaded workbook, Kindly check and try again!")

        logging.debug("Creating dataframe for the host details sheet dataframe")
        # Creating the host details sheet dataframe
        host_details_df = pd.read_excel(file_reader, sheet_name='Host Details')

        logging.info("Calling the pickle checker function")
        pickle_checker_result = pickle_checker(host_details_df=host_details_df)

        logging.info(f"Creating the checks for raising the host_details_df ==>\n{pickle_checker_result}\n")
        if pickle_checker_result[0] == 'Unsuccessful':
            logging.debug("Raising the Exception from the message gained in pickle_checker_result[1]")
            raise Exception(pickle_checker_result[1])

        # logging.info(f"Adding the path of the selected host details in the host_details_text_file in AppData/Local folder")
        # username = ((os.popen('cmd.exe /C "echo %username%"')).read()).strip()
        # host_details_file_text = rf"C:\Users\{username}\AppData\Local\CLI_Automation\Host_details_Pickle_file\Host_details_Path.txt"

        # with open(file=host_details_file_text, mode='w') as f:
        #     f.write(host_details_file_name)
        #     f.close()

        # del f

        logging.info(f'Read the host details\n\n{host_details_df.to_markdown()}\n')

        # Performing all the mandatory checks for the 'Host Details' sheet.
        unique_host_ips = host_details_df['Host_IP'].unique()

        # if(pd.NA in nan_test ):
        nan_test = any(pd.isna(element) for element in unique_host_ips)

        # if(np.nan in nan_test):
        #     nan_test = any(pd.isna(element) for element in unique_host_ips)            
        blank_Sr_no_list = []
        # If nan_test is true
        if nan_test:
            i = 0
            while i < len(host_details_df):
                if pd.isna(host_details_df.iloc[i]['Host_IP']):
                    blank_Sr_no_list.append(host_details_df.iloc[i]['Sr.No'])
                i += 1

        # Finding row with blank Vendor
        blank_Sr_no_list_for_Vendor = []
        unique_vendors_in_host_details = host_details_df['Vendor'].unique()

        logging.info("Checking for rows with blank Vendor details")
        nan_test = any(pd.isna(element) for element in unique_vendors_in_host_details)

        if nan_test:
            i = 0
            while i < len(host_details_df):
                if pd.isna(host_details_df.iloc[i]['Vendor']):
                    blank_Sr_no_list_for_Vendor.append(host_details_df.iloc[i]['Sr.No'])
                i += 1
        # temp_array = unique_host_ips[~np.isnan(unique_host_ips)]
        # temp_array = unique_host_ips[~pd.isna(unique_host_ips)]
        host_details_df.dropna(inplace=True)

        # Checking duplicated Host IPs in the Host Details sheet
        duplicate_host_ip_boolean_series = host_details_df.duplicated(subset=['Host_IP'], keep=False)
        duplicated_host_ip_Sr_no_list = []
        logging.info("Checking duplicated host ip details in the Host Details sheet")

        i = 0
        while i < duplicate_host_ip_boolean_series.size:
            if duplicate_host_ip_boolean_series[i]:
                temp_array_list = host_details_df[host_details_df['Host_IP'] == host_details_df.iloc[i]['Host_IP']]
                duplicated_host_ip_Sr_no_list.extend(list(temp_array_list['Sr.No']))
            i += 1

        if len(duplicated_host_ip_Sr_no_list) > 0:
            duplicated_host_ip_Sr_no_list = np.unique(np.array(duplicated_host_ip_Sr_no_list))

        if ((len(blank_Sr_no_list) > 0) and (len(duplicated_host_ip_Sr_no_list) > 0) and (
                len(blank_Sr_no_list_for_Vendor) > 0)):
            # file_reader.close()
            # del file_reader
            logging.error("Blank Host IP & Vendors and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                  f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}\n\nand \n\nBlank Vendor Details found for Sr no: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}\n\nand \n\nDuplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")

        if (len(blank_Sr_no_list) > 0) and (len(duplicated_host_ip_Sr_no_list) > 0):
            # file_reader.close()
            # del file_reader
            logging.error("Blank Host IP and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                  f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}\n\nand \n\nDuplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")

        if (len(blank_Sr_no_list) > 0) and (len(blank_Sr_no_list_for_Vendor) > 0):
            # file_reader.close()
            # del file_reader
            logging.error("Blank Host IP and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                  f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}\n\nand \n\nBlank Vendor Details found for Sr no: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}")

        if (len(blank_Sr_no_list_for_Vendor) > 0) and (len(duplicated_host_ip_Sr_no_list) > 0):
            # file_reader.close()
            # del file_reader
            logging.error("Blank Vendor and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                  f"Blank Vendor Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}\n\nand \n\nDuplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")

        if len(blank_Sr_no_list) > 0:
            # file_reader.close()
            # del file_reader
            logging.error("Blank Host IP details are found!")
            raise CustomException("Blank Host IP Details Found!",
                                  f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}")

        if len(duplicated_host_ip_Sr_no_list) > 0:
            # file_reader.close()
            # del file_reader
            logging.error("Duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                  f"Duplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")

        if len(blank_Sr_no_list_for_Vendor) > 0:
            # file_reader.close()
            # del file_reader
            logging.error("Blank Vendor details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                  f"Duplicate Host IPs found for Sr no: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}")

        blank_CR_ID_field_sr_no_list = [row['Sr.No'] for _, row in host_details_df.iterrows() if pd.isna(row['CR_ID'])]

        if len(blank_CR_ID_field_sr_no_list) > 0:
            logging.error(f"Blank CR IDs found for Below Sr.No. ====>\n{', '.join(blank_CR_ID_field_sr_no_list)}\n")
            raise CustomException("Blank CR IDs Found!",
                                  f"Blank CR IDs found for Below Sr.No. :\n{', '.join(blank_CR_ID_field_sr_no_list)}\n Kindly Check!")

        blank_host_name_field_sr_no_list = [row['Sr.No'] for _, row in host_details_df.iterrows() if
                                            pd.isna(row['Host_Name'])]
        if len(blank_host_name_field_sr_no_list) > 0:
            logging.error(
                f"Blank Host Names found for Below Sr.No. ====>\n{', '.join(blank_host_name_field_sr_no_list)}\n")
            raise CustomException("Blank Host Name/s Found!",
                                  f"Blank Host Name/s found for Below Sr.No. :\n{', '.join(blank_host_name_field_sr_no_list)}\n Kindly Check!")

        blank_vendor_field_sr_no_list = [row['Sr.No'] for _, row in host_details_df.iterrows() if
                                         pd.isna(row['Vendor'])]
        if len(blank_vendor_field_sr_no_list) > 0:
            logging.error(
                f"Blank Vendor field/s found for Below Sr.No. ====>\n{', '.join(blank_vendor_field_sr_no_list)}\n")
            raise CustomException("Blank Vendor field/s Found!",
                                  f"Blank Vendor field/s found for Below Sr.No. :\n{', '.join(blank_vendor_field_sr_no_list)}\n Kindly Check!")

        duplicated_host_names_sr_number_series = \
            (host_details_df[host_details_df.duplicated(subset=['Host_Name'], keep=False)])['Sr.No']
        if len(duplicated_host_names_sr_number_series) > 0:
            logging.error("Duplicated Host Name details are found!")
            raise CustomException("Host Name Details Incorrect!",
                                  f"Duplicate Host Names found for Sr no: {', '.join(str(element) for element in duplicated_host_names_sr_number_series)}")

        # getting the list of all the files present in the parent directory of host_details_file_name
        logging.info("Getting the list of all the files present in the parent directory of host_details_file_name")
        parent_dir = os.path.dirname(host_details_file_name)

        list_of_files_present_parent_dir = os.listdir(parent_dir)
        logging.debug(
            f"List of all the files present in the parent directory of host_details_file\n\n{'\n'.join(str(element) for element in list_of_files_present_parent_dir)}\n\n")

        missing_standard_input_design_template = []
        logging.info("Checking for the availability of Standard Input Design Templates")

        i = 0
        while i < unique_vendors_in_host_details.size:
            if unique_vendors_in_host_details[i].strip().upper() == 'NOKIA':
                if (not Path(os.path.join(os.path.join(os.path.dirname(parent_dir), 'Standard_Templates'),
                                          'Nokia_Design_Standard_Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Nokia')

            if unique_vendors_in_host_details[i].strip().upper() == 'ERICSSON':
                if (not Path(os.path.join(os.path.join(os.path.dirname(parent_dir), 'Standard_Templates'),
                                          'Ericsson_Design_Standard_Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Ericsson')

            if unique_vendors_in_host_details[i].strip().upper() == 'HUAWEI':
                if (not Path(os.path.join(os.path.join(os.path.dirname(parent_dir), 'Standard_Templates'),
                                          'Huawei_Design_Standard_Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Huawei')

            if unique_vendors_in_host_details[i].strip().upper() == 'CISCO':
                if (not Path(os.path.join(os.path.join(os.path.dirname(parent_dir), 'Standard_Templates'),
                                          'Cisco_Design_Standard_Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Cisco')
            i += 1

        if len(missing_standard_input_design_template) > 0:
            del host_details_file_name
            file_reader.close()
            del file_reader
            logging.error("Standard Input design template missing")
            raise CustomException("Standard Input Design Template Missing!",
                                  f"Standard Input design template missing for below mentioned Vendors:\n\n{', '.join(missing_standard_input_design_template)}")

        neo_parent_dir = os.path.join(parent_dir, "Design_Input_Sheets")
        Path(neo_parent_dir).mkdir(exist_ok=True, parents=True)

        logging.debug(f"Created new folder if not created earlier \'{parent_dir}\'\n")

        file_to_be_created = "{}_Design_Input_Sheet.xlsx"
        file_to_be_selected = "{}_Design_Standard_Input_Template.xlsx"

        # Task for creating the host ip sheets in the respective Vendor workbook
        thread_list_for_vendor_sheet_creation = []
        try:
            i = 0
            while i < unique_vendors_in_host_details.size:
                host_ips_sheets_required = host_details_df[
                    host_details_df['Vendor'].astype(str) == unique_vendors_in_host_details[i]]
                host_ips_sheets_required = np.array(host_ips_sheets_required['Host_IP'])

                # Checking the file existence for the input files for the user
                logging.debug(
                    f"Checking if the '{file_to_be_created.format(unique_vendors_in_host_details[i])}' exists or not, if yes then adding ip sheets, otherwise creating the file itself.")
                logging.debug(
                    f"Checking the condition working or not '=> not Path(os.path.join(neo_parent_dir,{file_to_be_created.format(unique_vendors_in_host_details[i])})).exists() ={(not Path(os.path.join(neo_parent_dir, file_to_be_created.format(unique_vendors_in_host_details[i]))).exists())}'")

                if ((not Path(os.path.join(neo_parent_dir,
                                           file_to_be_created.format(unique_vendors_in_host_details[i]))).exists())):
                    logging.debug(f"Creating {file_to_be_created.format(unique_vendors_in_host_details[i])}")
                    file_creater(
                        file=os.path.join(neo_parent_dir, file_to_be_created.format(unique_vendors_in_host_details[i])))

                logging.debug(
                    f"Checking for the presence of the 'Standard Template' worksheet in the {file_to_be_selected.format(unique_vendors_in_host_details[i])}")
                temp_reader = pd.ExcelFile(os.path.join(os.path.join(os.path.dirname(parent_dir), 'Standard_Templates'),
                                                        file_to_be_selected.format(unique_vendors_in_host_details[i])))

                if not 'Standard Template' in temp_reader.sheet_names:
                    del temp_reader
                    messagebox.showwarning("Template Sheet Missing!",
                                             f"'Standard Template' worksheet missing in {file_to_be_selected.format(unique_vendors_in_host_details[i])}, Kindly Check and Try Again!")
                    continue

                logging.info(
                    "Deleting the temp_reader (Created to find the existence of 'Standard Template' worksheet)")
                del temp_reader

                logging.debug(
                    f"Creating the input file with sheets for '{unique_vendors_in_host_details[i]}' via thread")

                # Creating Sheets for the unique IPs using threads
                thread = Thread(target=sheet_creater,
                                kwargs={'host_ips_sheets_required': host_ips_sheets_required,
                                        'file': os.path.join(neo_parent_dir, file_to_be_created.format(
                                            unique_vendors_in_host_details[i])),
                                        'standard_design_template_path': os.path.join(
                                            os.path.join(os.path.dirname(parent_dir), 'Standard_Templates'),
                                            file_to_be_selected.format(unique_vendors_in_host_details[i]))})
                thread_list_for_vendor_sheet_creation.append(thread)
                thread.daemon = True
                thread.start()
                i += 1

            logging.info("Stopping all the threads for vendor workbook after writing the information")
            i = 0
            while i < len(thread_list_for_vendor_sheet_creation):
                thread_list_for_vendor_sheet_creation[i].join()
                i += 1

        except CustomException as e:
            # global flag;
            flag = 'Unsuccessful'
            logging.error(
                f"{traceback.format_exc()}\n\nraised CustomException==>\ntitle = {e.title}\nmessage = {e.message}")

        except PermissionError as e:
            flag = 'Unsuccessful'
            logging.error(f"{traceback.format_exc()}\n\nException:==>\n\tTitle ===> {type(e)} \n\tMessage ==>{e}")
            messagebox.showerror("Permission Error!", str(e))

        except Exception as e:
            # global flag;
            flag = 'Unsuccessful'
            logging.error(f"{traceback.format_exc()}\n\nException:==>\n\tTitle ===> {type(e)} \n\tMessage ==>{e}")
            messagebox.showerror("Exception Occurred!", str(e))

        if flag != 'Unsuccessful':
            logging.debug(
                f"Invoking mpbn_node_login_file_creater method for \"MPBN_Node_Login.xlsx\" creation from \'host_details_df\' ==> \n{host_details_df.to_markdown()}\n")
            mpbn_node_login_file_creater(parent_dir=parent_dir,
                                         host_details_file_df=host_details_df)

        if flag != 'Unsuccessful':
            logging.info("Saving the selected host_file_name")
            file_path_saver(host_details_file_name)

            logging.info("Setting the flag status to 'Successful'")
            # messagebox.showinfo(title="Task Successfully Completed!",
            #                     message="Sheet Creater Task Completed Successfully!")
            flag = 'Successful'

    except CustomException as e:
        flag = 'Unsuccessful'
        logging.error(
            f"{traceback.format_exc()}\n\nraised CustomException==>\ntitle = {e.title}\nmessage = {e.message}")

    except Exception as e:
        flag = 'Unsuccessful'
        logging.error(f"{traceback.format_exc()}\n\nException:==>{e}")
        messagebox.showerror("Exception Occurred!", str(e))

    finally:
        del host_details_file_name
        file_reader.close()
        del file_reader

        logging.info(f"Returning Flag\n\n\t{flag=}")
        logging.shutdown()
        gc.collect()
        return flag

# print(main_func(file_name = r"C:\Users\emaienj\Downloads\VPLS_CLI_Design_Documents\VPLS_CLI_Design_Documents\Host_Details.xlsx"))